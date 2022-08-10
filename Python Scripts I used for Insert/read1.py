from openpyxl import load_workbook
import mysql.connector

book = load_workbook('C:\\Users\emrecinar\Desktop\modeo3_1.xlsx')
#print(book.sheetnames)

sheet = book.active

rows = sheet.rows
count = 0

headers = [cell.value for cell in next(rows)]
#(headers)
all_rows = []
for row in rows:
    if(count == 5):
        break
    data = []
    for title, cell in zip(headers, row):
        data.append(cell.value)

        if(title == "NOT"):
            break

    all_rows.append(data)
    count += 1
    
print(all_rows[0])

"""
conn = mysql.connector.connect(
    host = "localhost",
    user = "root",
    password = "",
    database = "modeo"
)

cur = conn.cursor()

our_list = ["sasaas","fdsfsfsd"]
#sql = "INSERT INTO cihazkimlik (Model, `Cihaz Seri No`, `Anakart No`, `UID No`, `Modem Kartı`, `Lcd Kartı`, `Şarj Kartı`, `Durumu`, `Modem Tipi`, `Müşteri Adı`, `Modem Seri Num 1`, `Modem Seri Num 2`, `Modem Seri Num 3`, `Modem Seri Num 4`, `Modem Seri Num 5`, `Modem Seri Num 6`, `Uretim Tarihi`, `Test Durumu`, `Notlar`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
sql = "INSERT INTO model (name) VALUES (%s)"
try:
    cur.executemany(sql, our_list)
    conn.commit()
except:
    print(conn.rollback())

cur.execute("SELECT * FROM model")

sekli = []
for x in cur: 
    sekli.append(x)
print(sekli)

cur.close()
conn.close()
"""
